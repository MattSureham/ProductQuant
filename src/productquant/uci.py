"""Pinned UCI Online Retail II acquisition and normalization."""

from __future__ import annotations

import ctypes
import errno
import hashlib
import json
import math
import os
import shutil
import stat
import subprocess
import sys
import tempfile
import urllib.error
import urllib.request
import uuid
import zipfile
from collections.abc import Iterator
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, BinaryIO

import duckdb
import openpyxl
import pyarrow as pa
import pyarrow.parquet as pq

from productquant.errors import IntegrityError, NetworkError, StateError

SOURCE_ID = "uci-online-retail-ii"
SOURCE_URL = (
    "https://cdn.uci-ics-mlr-prod.aws.uci.edu/502/"
    "online%2Bretail%2Bii.zip"
)
LICENCE_URL = "https://creativecommons.org/licenses/by/4.0/"
DATASET_URL = "https://archive.ics.uci.edu/dataset/502/online+retail+ii"
ARCHIVE_NAME = "online-retail-ii.zip"
ARCHIVE_BYTES = 45_622_418
ARCHIVE_SHA256 = "572e36277c2390fbfde10664750731e0a86f55e33470d91919085f0408e67bfb"
RANGE_CHUNK_BYTES = 4 * 1024 * 1024
MAX_RANGE_REQUESTS = 32
MEMBER_NAME = "online_retail_II.xlsx"
MEMBER_BYTES = 45_622_278
MEMBER_SHA256 = "bcbe73b35f5b7babf197fb0cb983a11f5d9ff929078d4aa53d171b1f2df2e980"
RAW_ARTIFACT_ID = f"{SOURCE_ID}:sha256:{ARCHIVE_SHA256}"
NORMALIZED_ARTIFACT_ID = f"{RAW_ARTIFACT_ID}:transaction-event.v1"
RAW_MANIFEST_SCHEMA = "productquant.raw-artifact-manifest.v1"
NORMALIZED_MANIFEST_SCHEMA = "productquant.normalized-manifest.v1"
RECEIPT_SCHEMA = "productquant.command-receipt.v1"
EVENT_SCHEMA_VERSION = "transaction_event.v1"
ADAPTER_VERSION = "uci-online-retail-ii.v1"
NORMALIZER_VERSION = "uci-online-retail-ii-normalizer.v1"
STITCH_RULE_ID = "uci-online-retail-ii-sheet-boundary-v1"
CUTOFF = datetime(2010, 12, 1)
SHEETS = ("Year 2009-2010", "Year 2010-2011")
SHEET_TOKENS = {
    "Year 2009-2010": "year-2009-2010",
    "Year 2010-2011": "year-2010-2011",
}
CAPABILITIES = {
    "transaction_event_history": "supported",
    "marketplace_listing_state": "unsupported",
    "marketplace_supply_competition": "unsupported",
    "independent_external_demand": "unsupported",
    "complete_product_opportunity_universe": "unsupported",
    "integrated_productquant_v0_1": "unsupported",
    "historical_provider_revision_state": "unknown",
    "timezone": "unknown",
}
HEADERS = (
    "Invoice",
    "StockCode",
    "Description",
    "Quantity",
    "InvoiceDate",
    "Price",
    "Customer ID",
    "Country",
)

EXPECTED_DATASET_ROW_COUNT = 1_044_848
EXPECTED_STITCH: dict[str, int] = {
    "first_sheet_input": 525_461,
    "first_sheet_retained": 502_938,
    "second_sheet_input": 541_910,
    "second_sheet_retained": 541_910,
    "excluded_cross_sheet_rows": 22_523,
}
EXPECTED_PROFILE: dict[str, Any] = {
    "event_time_min": "2009-12-01T07:45:00.000000",
    "event_time_max": "2011-12-09T12:50:00.000000",
    "unique_event_ids": 1_044_848,
    "unique_stock_codes": 5_305,
    "unique_invoices": 53_628,
    "cancellation_rows": 19_165,
    "negative_quantity_rows": 22_557,
    "zero_price_rows": 6_024,
    "negative_price_rows": 5,
    "missing_description_rows": 4_275,
    "missing_customer_rows": 235_287,
    "numeric_description_rows": 4,
    "quantity_min": -80_995,
    "quantity_max": 80_995,
    "unit_price_min": "-53594.360",
    "unit_price_max": "38970.000",
}

PARQUET_COMPRESSION = "zstd"
PARQUET_FORMAT_VERSION = "2.6"
PARQUET_ROW_GROUP_SIZE = 25_000
PARQUET_USE_DICTIONARY = True
PARQUET_WRITE_STATISTICS = True

EVENT_SCHEMA = pa.schema(
    [
        pa.field("event_id", pa.string(), nullable=False),
        pa.field("raw_artifact_id", pa.string(), nullable=False),
        pa.field("source_id", pa.string(), nullable=False),
        pa.field("source_member", pa.string(), nullable=False),
        pa.field("source_sheet", pa.string(), nullable=False),
        pa.field("source_row_number", pa.int64(), nullable=False),
        pa.field("event_time_local", pa.timestamp("us"), nullable=False),
        pa.field("source_invoice_id", pa.string(), nullable=False),
        pa.field("source_product_code", pa.string(), nullable=False),
        pa.field("description_observed", pa.string(), nullable=True),
        pa.field("quantity", pa.int64(), nullable=False),
        pa.field("unit_price", pa.decimal128(18, 3), nullable=False),
        pa.field("currency_code", pa.string(), nullable=False),
        pa.field("customer_reference", pa.string(), nullable=True),
        pa.field("country_observed", pa.string(), nullable=False),
        pa.field("invoice_is_cancellation", pa.bool_(), nullable=False),
        pa.field("schema_version", pa.string(), nullable=False),
    ]
)


def utc_now() -> str:
    return datetime.now(UTC).isoformat(timespec="microseconds").replace("+00:00", "Z")


def _json_bytes(value: dict[str, Any]) -> bytes:
    return (json.dumps(value, sort_keys=True, separators=(",", ":")) + "\n").encode()


def _chmod_private(path: Path, *, directory: bool = False) -> None:
    if os.name == "posix":
        path.chmod(stat.S_IRWXU if directory else stat.S_IRUSR | stat.S_IWUSR)


def _path_entry_exists(path: Path) -> bool:
    """Return whether a directory entry exists without following symlinks."""
    try:
        path.lstat()
    except FileNotFoundError:
        return False
    return True


def _require_real_directory(path: Path, label: str) -> os.stat_result:
    try:
        result = path.lstat()
    except FileNotFoundError as exc:
        raise IntegrityError(f"{label} is missing: {path}") from exc
    if not stat.S_ISDIR(result.st_mode):
        raise IntegrityError(f"{label} is not a real directory: {path}")
    return result


def _require_regular_file(path: Path, label: str) -> os.stat_result:
    try:
        result = path.lstat()
    except FileNotFoundError as exc:
        raise IntegrityError(f"{label} is missing: {path}") from exc
    if not stat.S_ISREG(result.st_mode):
        raise IntegrityError(f"{label} is not a real regular file: {path}")
    return result


def _containing_git_worktree(path: Path) -> Path | None:
    for candidate in (path, *path.parents):
        if _path_entry_exists(candidate / ".git"):
            return candidate
    return None


def _assert_data_root_policy(data_root: Path) -> None:
    """Require customer-bearing state inside a worktree to be Git-ignored."""
    root = data_root.resolve()
    repository = _containing_git_worktree(root)
    if repository is None:
        return
    try:
        result = subprocess.run(
            [
                "git",
                "-C",
                str(repository),
                "check-ignore",
                "--no-index",
                "-q",
                "--",
                f"{root}{os.sep}",
            ],
            check=False,
            capture_output=True,
        )
    except OSError as exc:
        raise StateError(
            "cannot verify that the in-worktree data root is Git-ignored"
        ) from exc
    if result.returncode != 0:
        raise StateError("in-worktree data root is not fully Git-ignored")


def ensure_private_directory(path: Path) -> None:
    """Create missing directory components privately; preserve existing modes."""
    missing: list[Path] = []
    cursor = path
    while not _path_entry_exists(cursor):
        missing.append(cursor)
        parent = cursor.parent
        if parent == cursor:  # pragma: no cover - an existing filesystem root is normal
            raise StateError("cannot find an existing parent for the data root")
        cursor = parent
    try:
        existing = cursor.lstat()
    except OSError as exc:  # pragma: no cover - raced filesystem state
        raise StateError("cannot inspect an existing data-root path") from exc
    if not stat.S_ISDIR(existing.st_mode):
        raise StateError("an existing data-root path is not a directory")

    for directory in reversed(missing):
        try:
            directory.mkdir(mode=0o700)
        except FileExistsError:
            try:
                appeared = directory.lstat()
            except OSError as exc:  # pragma: no cover - raced filesystem state
                raise StateError("cannot inspect an appeared data-root path") from exc
            if not stat.S_ISDIR(appeared.st_mode):
                raise StateError("a non-directory appeared in the data-root path")
        else:
            _chmod_private(directory, directory=True)


def ensure_private_data_tree(data_root: Path, target: Path) -> None:
    """Create private missing paths without changing any existing directory."""
    root = data_root.resolve()
    destination = target.resolve()
    try:
        relative = destination.relative_to(root)
    except ValueError as exc:  # pragma: no cover - internal path invariant
        raise StateError("artifact target escapes the data root") from exc
    _assert_data_root_policy(root)
    ensure_private_directory(root)
    current = root
    for component in relative.parts:
        current = current / component
        ensure_private_directory(current)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def fsync_file(path: Path) -> None:
    with path.open("rb") as handle:
        os.fsync(handle.fileno())


def fsync_directory(path: Path) -> None:
    if os.name == "posix":
        descriptor = os.open(path, os.O_RDONLY)
        try:
            os.fsync(descriptor)
        finally:
            os.close(descriptor)


def _rename_noreplace(source: Path, destination: Path) -> None:
    """Atomically rename one sibling path without replacing any destination."""
    if source.parent.resolve() != destination.parent.resolve():
        raise StateError("atomic publication requires sibling paths")
    source_bytes = os.fsencode(source)
    destination_bytes = os.fsencode(destination)
    try:
        if sys.platform == "darwin":
            library = ctypes.CDLL(None, use_errno=True)
            ctypes.set_errno(0)
            operation = library.renamex_np
            operation.argtypes = [ctypes.c_char_p, ctypes.c_char_p, ctypes.c_uint]
            operation.restype = ctypes.c_int
            result = operation(source_bytes, destination_bytes, 0x00000004)
        elif sys.platform.startswith("linux"):
            library = ctypes.CDLL(None, use_errno=True)
            ctypes.set_errno(0)
            operation = library.renameat2
            operation.argtypes = [
                ctypes.c_int,
                ctypes.c_char_p,
                ctypes.c_int,
                ctypes.c_char_p,
                ctypes.c_uint,
            ]
            operation.restype = ctypes.c_int
            result = operation(-100, source_bytes, -100, destination_bytes, 1)
        elif sys.platform == "win32":  # pragma: no cover - CI host is POSIX
            library = ctypes.WinDLL("kernel32", use_last_error=True)
            operation = library.MoveFileExW
            operation.argtypes = [
                ctypes.c_wchar_p,
                ctypes.c_wchar_p,
                ctypes.c_ulong,
            ]
            operation.restype = ctypes.c_int
            ctypes.set_last_error(0)
            result = operation(str(source), str(destination), 0)
        else:
            raise StateError("atomic no-replace publication is unsupported on this host")
    except AttributeError as exc:
        raise StateError("atomic no-replace publication is unavailable on this host") from exc

    if result == 0:
        return
    error_number = (
        ctypes.get_last_error() if sys.platform == "win32" else ctypes.get_errno()
    )
    collision_errors = {errno.EEXIST, errno.ENOTEMPTY, errno.EISDIR}
    if sys.platform == "win32":  # pragma: no cover - CI host is POSIX
        collision_errors.update({80, 183})
    if error_number in collision_errors:
        raise StateError("artifact target appeared during publication")
    unsupported = {
        errno.ENOSYS,
        errno.EINVAL,
        getattr(errno, "ENOTSUP", errno.EINVAL),
        getattr(errno, "EOPNOTSUPP", errno.EINVAL),
    }
    if error_number in unsupported:
        raise StateError("atomic no-replace publication is unsupported here")
    raise StateError(
        f"atomic artifact publication failed: errno={error_number}"
    )


def _publish_noreplace(source: Path, destination: Path, *, directory: bool) -> None:
    try:
        source_stat = source.lstat()
    except FileNotFoundError as exc:
        raise StateError("publication source is missing") from exc
    expected_type = stat.S_ISDIR if directory else stat.S_ISREG
    if not expected_type(source_stat.st_mode):
        raise StateError("publication source has an unsupported filesystem type")
    _rename_noreplace(source, destination)
    try:
        fsync_directory(destination.parent)
    except OSError as exc:
        raise StateError("artifact publication directory sync failed") from exc


def atomic_publish_directory(stage: Path, destination: Path) -> None:
    """Publish a complete real directory atomically without clobbering a target."""
    _publish_noreplace(stage, destination, directory=True)


def raw_directory(data_root: Path) -> Path:
    return data_root / "raw" / SOURCE_ID / ARCHIVE_SHA256


def normalized_directory(data_root: Path) -> Path:
    return (
        data_root
        / "normalized"
        / SOURCE_ID
        / ARCHIVE_SHA256
        / "transaction-event-v1"
    )


def receipts_directory(data_root: Path) -> Path:
    return data_root / "receipts" / SOURCE_ID


def _load_json(path: Path) -> dict[str, Any]:
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise IntegrityError(f"invalid JSON artifact: {path}") from exc
    if not isinstance(value, dict):
        raise IntegrityError(f"JSON artifact is not an object: {path}")
    return value


def _require_keys(value: dict[str, Any], expected: set[str], label: str) -> None:
    if set(value) != expected:
        raise IntegrityError(
            f"{label} keys differ: expected {sorted(expected)}, found {sorted(value)}"
        )


def _require_utc_timestamp(value: Any, label: str) -> None:
    if not isinstance(value, str) or not value.endswith("Z"):
        raise IntegrityError(f"{label} is not a UTC timestamp")
    if len(value) != 27 or value[10] != "T" or value[19] != ".":
        raise IntegrityError(f"{label} is not a microsecond UTC timestamp")
    try:
        parsed = datetime.fromisoformat(value.removesuffix("Z") + "+00:00")
    except ValueError as exc:
        raise IntegrityError(f"{label} is not a UTC timestamp") from exc
    if parsed.tzinfo != UTC:
        raise IntegrityError(f"{label} is not a UTC timestamp")


def _assert_exact_bundle(directory: Path, expected_names: set[str]) -> None:
    _require_real_directory(directory, "artifact bundle")
    actual = {entry.name for entry in directory.iterdir()}
    if actual != expected_names:
        raise IntegrityError(
            f"artifact bundle contents differ: expected {sorted(expected_names)}, "
            f"found {sorted(actual)}"
        )


def _verify_private_mode(path: Path, *, directory: bool = False) -> None:
    if os.name != "posix":
        return
    expected = 0o700 if directory else 0o600
    actual = stat.S_IMODE(path.lstat().st_mode)
    if actual != expected:
        raise IntegrityError(
            f"artifact permission mode differs for {path.name}: "
            f"expected {oct(expected)}, found {oct(actual)}"
        )


def _verify_archive(path: Path) -> dict[str, Any]:
    archive_stat = _require_regular_file(path, "raw archive")
    size = archive_stat.st_size
    if size != ARCHIVE_BYTES:
        raise IntegrityError(f"archive byte count differs: {size}")
    digest = sha256_file(path)
    if digest != ARCHIVE_SHA256:
        raise IntegrityError(f"archive SHA-256 differs: {digest}")
    try:
        with zipfile.ZipFile(path) as archive:
            members = archive.infolist()
            if len(members) != 1 or members[0].filename != MEMBER_NAME:
                raise IntegrityError("archive member set differs")
            if members[0].file_size != MEMBER_BYTES:
                raise IntegrityError("workbook byte count differs")
            member_digest = hashlib.sha256()
            with archive.open(members[0]) as handle:
                for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                    member_digest.update(chunk)
            if member_digest.hexdigest() != MEMBER_SHA256:
                raise IntegrityError("workbook SHA-256 differs")
    except zipfile.BadZipFile as exc:
        raise IntegrityError("archive is not a valid ZIP") from exc
    return {"bytes": size, "sha256": digest}


def _verify_raw_directory(directory: Path) -> dict[str, Any]:
    _assert_exact_bundle(directory, {ARCHIVE_NAME, "manifest.json"})
    _verify_private_mode(directory, directory=True)
    archive_path = directory / ARCHIVE_NAME
    manifest_path = directory / "manifest.json"
    _require_regular_file(manifest_path, "raw manifest")
    _verify_private_mode(archive_path)
    _verify_private_mode(manifest_path)
    _verify_archive(archive_path)
    manifest = _load_json(manifest_path)
    _require_keys(
        manifest,
        {
            "schema_version",
            "artifact_id",
            "source",
            "retrieval",
            "archive",
            "member",
            "adapter_version",
            "raw_response_reference",
            "capabilities",
        },
        "raw manifest",
    )
    expected = {
        "schema_version": RAW_MANIFEST_SCHEMA,
        "artifact_id": RAW_ARTIFACT_ID,
        "archive": {
            "path": ARCHIVE_NAME,
            "bytes": ARCHIVE_BYTES,
            "sha256": ARCHIVE_SHA256,
        },
        "member": {
            "name": MEMBER_NAME,
            "bytes": MEMBER_BYTES,
            "sha256": MEMBER_SHA256,
        },
    }
    for key, value in expected.items():
        if manifest.get(key) != value:
            raise IntegrityError(f"raw manifest field differs: {key}")
    source = manifest.get("source")
    if not isinstance(source, dict):
        raise IntegrityError("raw source manifest is missing")
    _require_keys(
        source,
        {
            "id",
            "original_url",
            "final_url",
            "licence_spdx",
            "licence_url",
            "attribution",
        },
        "raw source manifest",
    )
    if source.get("id") != SOURCE_ID or source.get("original_url") != SOURCE_URL:
        raise IntegrityError("raw source identity differs")
    if not isinstance(source.get("final_url"), str) or not source["final_url"].startswith(
        "https://"
    ):
        raise IntegrityError("raw final URL differs")
    if source.get("licence_spdx") != "CC-BY-4.0" or source.get(
        "licence_url"
    ) != LICENCE_URL:
        raise IntegrityError("raw licence declaration differs")
    if not isinstance(source.get("attribution"), str) or "10.24432/C5CG6D" not in source[
        "attribution"
    ]:
        raise IntegrityError("raw attribution differs")
    retrieval = manifest.get("retrieval")
    if not isinstance(retrieval, dict):
        raise IntegrityError("raw retrieval manifest is missing")
    _require_keys(
        retrieval,
        {
            "retrieved_at_utc",
            "mode",
            "http_status",
            "etag",
            "last_modified",
            "content_type",
            "content_length",
            "rate_limit_state",
            "query_parameters",
        },
        "raw retrieval manifest",
    )
    _require_utc_timestamp(retrieval.get("retrieved_at_utc"), "retrieved_at_utc")
    if retrieval.get("mode") != "download" or retrieval.get("http_status") not in {
        200,
        206,
    }:
        raise IntegrityError("raw retrieval mode/status differs")
    if retrieval.get("rate_limit_state") != "unknown" or retrieval.get(
        "query_parameters"
    ) != {}:
        raise IntegrityError("raw retrieval query/rate-limit state differs")
    for header in ("etag", "last_modified", "content_type", "content_length"):
        if retrieval.get(header) is not None and not isinstance(retrieval[header], str):
            raise IntegrityError(f"raw retrieval header type differs: {header}")
    if manifest.get("adapter_version") != ADAPTER_VERSION:
        raise IntegrityError("raw adapter version differs")
    if manifest.get("raw_response_reference") != ARCHIVE_NAME:
        raise IntegrityError("raw response reference differs")
    if manifest.get("capabilities") != CAPABILITIES:
        raise IntegrityError("raw capability declarations differ")
    return manifest


def verify_raw_bundle(data_root: Path) -> dict[str, Any]:
    return _verify_raw_directory(raw_directory(data_root))


def _copy_response(response: BinaryIO, target: Path) -> tuple[int, str]:
    digest = hashlib.sha256()
    size = 0
    with target.open("xb") as output:
        _chmod_private(target)
        while True:
            remaining_with_sentinel = ARCHIVE_BYTES + 1 - size
            if remaining_with_sentinel <= 0:
                raise IntegrityError("downloaded archive exceeds pinned byte count")
            try:
                chunk = response.read(min(1024 * 1024, remaining_with_sentinel))
            except (urllib.error.URLError, TimeoutError, OSError) as exc:
                raise NetworkError(
                    f"UCI archive download interrupted: {type(exc).__name__}"
                ) from exc
            if not chunk:
                break
            if len(chunk) > remaining_with_sentinel:
                chunk = chunk[:remaining_with_sentinel]
            output.write(chunk)
            digest.update(chunk)
            size += len(chunk)
            if size > ARCHIVE_BYTES:
                output.flush()
                os.fsync(output.fileno())
                raise IntegrityError("downloaded archive exceeds pinned byte count")
        output.flush()
        os.fsync(output.fileno())
    return size, digest.hexdigest()


def _resume_short_download(target: Path) -> None:
    """Resume a cleanly truncated full response with bounded HTTP range requests."""
    requests = 0
    while target.stat().st_size < ARCHIVE_BYTES:
        if requests >= MAX_RANGE_REQUESTS:
            raise NetworkError("UCI archive download remained incomplete after retries")
        start = target.stat().st_size
        end = min(start + RANGE_CHUNK_BYTES, ARCHIVE_BYTES) - 1
        request = urllib.request.Request(
            SOURCE_URL,
            headers={
                "User-Agent": "ProductQuant/0.1 UCI-source-adapter",
                "Range": f"bytes={start}-{end}",
            },
        )
        try:
            response = urllib.request.urlopen(request, timeout=60)
        except (urllib.error.URLError, TimeoutError, OSError) as exc:
            raise NetworkError(
                f"UCI archive range download failed: {type(exc).__name__}"
            ) from exc
        with response:
            status = getattr(response, "status", None)
            content_range = response.headers.get("Content-Range")
            expected_range = f"bytes {start}-{end}/{ARCHIVE_BYTES}"
            if status != 206 or content_range != expected_range:
                raise IntegrityError("UCI range response metadata differs")
            written = 0
            with target.open("ab") as output:
                while True:
                    try:
                        chunk = response.read(1024 * 1024)
                    except (urllib.error.URLError, TimeoutError, OSError) as exc:
                        raise NetworkError(
                            f"UCI archive range download interrupted: {type(exc).__name__}"
                        ) from exc
                    if not chunk:
                        break
                    output.write(chunk)
                    written += len(chunk)
                    if written > end - start + 1:
                        raise IntegrityError("UCI range response exceeded its requested size")
                output.flush()
                os.fsync(output.fileno())
        if written == 0:
            raise NetworkError("UCI archive range download returned no bytes")
        requests += 1


def fetch(data_root: Path, *, offline: bool = False) -> dict[str, Any]:
    data_root = data_root.resolve()
    destination = raw_directory(data_root)
    if _path_entry_exists(destination):
        manifest = verify_raw_bundle(data_root)
        return _artifact_summary("raw", destination, manifest, "verified")
    if offline:
        raise StateError("offline mode requires a complete valid raw bundle")

    parent = destination.parent
    ensure_private_data_tree(data_root, parent)
    stage = Path(tempfile.mkdtemp(prefix=f".{ARCHIVE_SHA256}.part-", dir=parent))
    _chmod_private(stage, directory=True)
    archive_path = stage / ARCHIVE_NAME
    try:
        request = urllib.request.Request(
            SOURCE_URL,
            headers={"User-Agent": "ProductQuant/0.1 UCI-source-adapter"},
        )
        try:
            response = urllib.request.urlopen(request, timeout=60)
        except (urllib.error.URLError, TimeoutError, OSError) as exc:
            raise NetworkError(f"UCI archive download failed: {type(exc).__name__}") from exc
        with response:
            size, digest = _copy_response(response, archive_path)
            headers = response.headers
            final_url = response.geturl()
            status = getattr(response, "status", 200)
        if size < ARCHIVE_BYTES:
            _resume_short_download(archive_path)
            size = archive_path.stat().st_size
            digest = sha256_file(archive_path)
        if size != ARCHIVE_BYTES or digest != ARCHIVE_SHA256:
            raise IntegrityError(
                f"downloaded archive differs: bytes={size}, sha256={digest}"
            )
        manifest = {
            "schema_version": RAW_MANIFEST_SCHEMA,
            "artifact_id": RAW_ARTIFACT_ID,
            "source": {
                "id": SOURCE_ID,
                "original_url": SOURCE_URL,
                "final_url": final_url,
                "licence_spdx": "CC-BY-4.0",
                "licence_url": LICENCE_URL,
                "attribution": (
                    "Chen, D. (2012). Online Retail II [Dataset]. UCI Machine "
                    "Learning Repository. https://doi.org/10.24432/C5CG6D"
                ),
            },
            "retrieval": {
                "retrieved_at_utc": utc_now(),
                "mode": "download",
                "http_status": status,
                "etag": headers.get("ETag"),
                "last_modified": headers.get("Last-Modified"),
                "content_type": headers.get("Content-Type"),
                "content_length": headers.get("Content-Length"),
                "rate_limit_state": "unknown",
                "query_parameters": {},
            },
            "archive": {
                "path": ARCHIVE_NAME,
                "bytes": ARCHIVE_BYTES,
                "sha256": ARCHIVE_SHA256,
            },
            "member": {
                "name": MEMBER_NAME,
                "bytes": MEMBER_BYTES,
                "sha256": MEMBER_SHA256,
            },
            "adapter_version": ADAPTER_VERSION,
            "raw_response_reference": ARCHIVE_NAME,
            "capabilities": CAPABILITIES,
        }
        manifest_path = stage / "manifest.json"
        manifest_path.write_bytes(_json_bytes(manifest))
        _chmod_private(manifest_path)
        fsync_file(archive_path)
        fsync_file(manifest_path)
        fsync_directory(stage)
        _verify_raw_directory(stage)
        atomic_publish_directory(stage, destination)
        return _artifact_summary("raw", destination, manifest, "created")
    except Exception:
        shutil.rmtree(stage, ignore_errors=True)
        raise


def _artifact_summary(
    kind: str, directory: Path, manifest: dict[str, Any], action: str
) -> dict[str, Any]:
    data_name = ARCHIVE_NAME if kind == "raw" else "events.parquet"
    return {
        "artifact_id": manifest["artifact_id"],
        "manifest_path": str((directory / "manifest.json").resolve()),
        "data_path": str((directory / data_name).resolve()),
        "action": action,
    }


def _extract_workbook(archive_path: Path, target: Path) -> None:
    with zipfile.ZipFile(archive_path) as archive:
        with archive.open(MEMBER_NAME) as source, target.open("xb") as output:
            _chmod_private(target)
            shutil.copyfileobj(source, output, length=1024 * 1024)
            output.flush()
            os.fsync(output.fileno())
    if target.stat().st_size != MEMBER_BYTES or sha256_file(target) != MEMBER_SHA256:
        raise IntegrityError("transient workbook differs from pinned member")


def _as_identifier(value: Any, field: str) -> str:
    if isinstance(value, bool) or value is None:
        raise IntegrityError(f"{field} is not a supported identifier")
    if isinstance(value, str):
        return value
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and math.isfinite(value) and value.is_integer():
        return str(int(value))
    raise IntegrityError(f"{field} is not an integral identifier or string")


def _as_description(value: Any) -> tuple[str | None, bool]:
    if value is None:
        return None, False
    if isinstance(value, str):
        return value, False
    if isinstance(value, bool):
        raise IntegrityError("Description has an unsupported cell type")
    if isinstance(value, int):
        return str(value), True
    if isinstance(value, float) and math.isfinite(value) and value.is_integer():
        return str(int(value)), True
    raise IntegrityError("Description has an unsupported cell type")


def _as_quantity(value: Any) -> int:
    if isinstance(value, bool) or value is None:
        raise IntegrityError("Quantity is not an integer")
    if isinstance(value, int):
        quantity = value
    elif isinstance(value, float) and math.isfinite(value) and value.is_integer():
        quantity = int(value)
    else:
        raise IntegrityError("Quantity is not an integer")
    if quantity < -(1 << 63) or quantity > (1 << 63) - 1:
        raise IntegrityError("Quantity does not fit int64")
    return quantity


def _as_price(value: Any) -> Decimal:
    if isinstance(value, bool) or value is None or not isinstance(value, (int, float, Decimal)):
        raise IntegrityError("Price is not numeric")
    try:
        price = Decimal(str(value))
    except InvalidOperation as exc:
        raise IntegrityError("Price is not a valid decimal") from exc
    if not price.is_finite():
        raise IntegrityError("Price is not a finite decimal")
    scale = max(0, -price.as_tuple().exponent)
    if scale > 3:
        raise IntegrityError("Price precision exceeds decimal128(18,3)")
    try:
        quantized = price.quantize(Decimal("0.001"))
    except InvalidOperation as exc:
        raise IntegrityError("Price does not fit decimal128(18,3)") from exc
    if quantized != price or len(quantized.as_tuple().digits) > 18:
        raise IntegrityError("Price does not fit decimal128(18,3)")
    return quantized


def _row_to_event(
    sheet: str, row_number: int, row: tuple[Any, ...]
) -> tuple[dict[str, Any], bool]:
    if len(row) != len(HEADERS):
        raise IntegrityError(f"source row width differs at {sheet}:{row_number}")
    invoice, stock, description, quantity, event_time, price, customer, country = row
    if not isinstance(event_time, datetime) or event_time.tzinfo is not None:
        raise IntegrityError(f"InvoiceDate is not a naïve datetime at {sheet}:{row_number}")
    if country is None or not isinstance(country, str):
        raise IntegrityError(f"Country is not a required string at {sheet}:{row_number}")
    invoice_id = _as_identifier(invoice, "Invoice")
    stock_code = _as_identifier(stock, "StockCode")
    description_value, numeric_description = _as_description(description)
    customer_value = None if customer is None else _as_identifier(customer, "Customer ID")
    event_id = f"{RAW_ARTIFACT_ID}:{SHEET_TOKENS[sheet]}:{row_number}"
    return (
        {
            "event_id": event_id,
            "raw_artifact_id": RAW_ARTIFACT_ID,
            "source_id": SOURCE_ID,
            "source_member": MEMBER_NAME,
            "source_sheet": sheet,
            "source_row_number": row_number,
            "event_time_local": event_time,
            "source_invoice_id": invoice_id,
            "source_product_code": stock_code,
            "description_observed": description_value,
            "quantity": _as_quantity(quantity),
            "unit_price": _as_price(price),
            "currency_code": "GBP",
            "customer_reference": customer_value,
            "country_observed": country,
            "invoice_is_cancellation": invoice_id.startswith("C"),
            "schema_version": EVENT_SCHEMA_VERSION,
        },
        numeric_description,
    )


def iter_workbook_events(
    workbook_path: Path,
) -> Iterator[tuple[dict[str, Any], bool]]:
    try:
        workbook = openpyxl.load_workbook(
            workbook_path, read_only=True, data_only=False
        )
    except Exception as exc:
        raise IntegrityError("workbook cannot be opened") from exc
    try:
        if tuple(workbook.sheetnames) != SHEETS:
            raise IntegrityError(f"workbook sheet set/order differs: {workbook.sheetnames}")
        for sheet_name in SHEETS:
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(values_only=False)
            try:
                header = tuple(cell.value for cell in next(rows))
            except StopIteration as exc:
                raise IntegrityError(f"sheet is empty: {sheet_name}") from exc
            if header != HEADERS:
                raise IntegrityError(f"workbook header differs in {sheet_name}")
            for row_number, cells in enumerate(rows, start=2):
                if any(cell.data_type in {"f", "e"} for cell in cells):
                    raise IntegrityError(
                        f"workbook formula or error cell is unsupported at "
                        f"{sheet_name}:{row_number}"
                    )
                row = tuple(cell.value for cell in cells)
                event, numeric_description = _row_to_event(sheet_name, row_number, row)
                event_time = event["event_time_local"]
                retain = (
                    sheet_name == SHEETS[0] and event_time < CUTOFF
                ) or (
                    sheet_name == SHEETS[1] and event_time >= CUTOFF
                )
                if retain:
                    yield event, numeric_description
    finally:
        workbook.close()


def _git_state() -> tuple[str | None, bool | None]:
    repository = next(
        (
            parent
            for parent in Path(__file__).resolve().parents
            if (parent / ".git").exists()
        ),
        None,
    )
    if repository is None:
        return None, None
    try:
        revision = subprocess.run(
            ["git", "-C", str(repository), "rev-parse", "HEAD"],
            check=True,
            capture_output=True,
            text=True,
        ).stdout.strip()
        dirty = bool(
            subprocess.run(
                ["git", "-C", str(repository), "status", "--porcelain"],
                check=True,
                capture_output=True,
                text=True,
            ).stdout
        )
        return revision, dirty
    except (OSError, subprocess.SubprocessError):
        return None, None


def _dependency_versions() -> dict[str, str]:
    return {
        "duckdb": duckdb.__version__,
        "openpyxl": openpyxl.__version__,
        "pyarrow": pa.__version__,
    }


def _profile_parquet(path: Path) -> dict[str, Any]:
    connection = duckdb.connect(":memory:")
    try:
        row = connection.execute(
            """
            SELECT
              count(*)::BIGINT,
              min(event_time_local), max(event_time_local),
              count(DISTINCT event_id)::BIGINT,
              count(DISTINCT source_product_code)::BIGINT,
              count(DISTINCT source_invoice_id)::BIGINT,
              count(*) FILTER (WHERE invoice_is_cancellation)::BIGINT,
              count(*) FILTER (WHERE quantity < 0)::BIGINT,
              count(*) FILTER (WHERE unit_price = 0)::BIGINT,
              count(*) FILTER (WHERE unit_price < 0)::BIGINT,
              count(*) FILTER (WHERE description_observed IS NULL)::BIGINT,
              count(*) FILTER (WHERE customer_reference IS NULL)::BIGINT,
              min(quantity), max(quantity), min(unit_price), max(unit_price),
              count(*) FILTER (
                WHERE source_sheet = 'Year 2009-2010'
              )::BIGINT,
              count(*) FILTER (
                WHERE source_sheet = 'Year 2010-2011'
              )::BIGINT,
              count(*) FILTER (
                WHERE source_sheet = 'Year 2009-2010'
                  AND event_time_local >= TIMESTAMP '2010-12-01'
              )::BIGINT,
              count(*) FILTER (
                WHERE source_sheet = 'Year 2010-2011'
                  AND event_time_local < TIMESTAMP '2010-12-01'
              )::BIGINT
            FROM read_parquet(?)
            """,
            [str(path)],
        ).fetchone()
    finally:
        connection.close()
    assert row is not None
    return {
        "row_count": row[0],
        "event_time_min": row[1].isoformat(timespec="microseconds"),
        "event_time_max": row[2].isoformat(timespec="microseconds"),
        "unique_event_ids": row[3],
        "unique_stock_codes": row[4],
        "unique_invoices": row[5],
        "cancellation_rows": row[6],
        "negative_quantity_rows": row[7],
        "zero_price_rows": row[8],
        "negative_price_rows": row[9],
        "missing_description_rows": row[10],
        "missing_customer_rows": row[11],
        "quantity_min": row[12],
        "quantity_max": row[13],
        "unit_price_min": f"{row[14]:.3f}",
        "unit_price_max": f"{row[15]:.3f}",
        "first_sheet_retained": row[16],
        "second_sheet_retained": row[17],
        "first_sheet_boundary_violations": row[18],
        "second_sheet_boundary_violations": row[19],
    }


def _assert_profile(profile: dict[str, Any], manifest_profile: dict[str, Any]) -> None:
    for key, expected in EXPECTED_PROFILE.items():
        if key == "numeric_description_rows":
            actual = manifest_profile.get(key)
        else:
            actual = profile.get(key)
        if actual != expected:
            raise IntegrityError(
                f"normalized full-data invariant differs for {key}: "
                f"expected {expected!r}, found {actual!r}"
            )
    if profile["first_sheet_boundary_violations"] != 0:
        raise IntegrityError("first-sheet cutoff invariant differs")
    if profile["second_sheet_boundary_violations"] != 0:
        raise IntegrityError("second-sheet cutoff invariant differs")


def _assert_dataset_and_stitch_profile(profile: dict[str, Any]) -> None:
    if profile["row_count"] != EXPECTED_DATASET_ROW_COUNT:
        raise IntegrityError(
            "normalized full-data invariant differs for row_count: "
            f"expected {EXPECTED_DATASET_ROW_COUNT!r}, found {profile['row_count']!r}"
        )
    for key in ("first_sheet_retained", "second_sheet_retained"):
        if profile[key] != EXPECTED_STITCH[key]:
            raise IntegrityError(
                f"normalized stitch invariant differs for {key}: "
                f"expected {EXPECTED_STITCH[key]!r}, found {profile[key]!r}"
            )


def _schema_text(schema: pa.Schema) -> str:
    return schema.to_string(show_field_metadata=True, show_schema_metadata=True)


def normalize(data_root: Path) -> dict[str, Any]:
    data_root = data_root.resolve()
    raw_manifest = verify_raw_bundle(data_root)
    destination = normalized_directory(data_root)
    if _path_entry_exists(destination):
        manifest = verify_normalized_bundle(data_root)
        return _artifact_summary("normalized", destination, manifest, "verified")

    parent = destination.parent
    ensure_private_data_tree(data_root, parent)
    stage = Path(tempfile.mkdtemp(prefix=".transaction-event-v1.part-", dir=parent))
    _chmod_private(stage, directory=True)
    parquet_path = stage / "events.parquet"
    workspace = Path(tempfile.mkdtemp(prefix="productquant-uci-"))
    _chmod_private(workspace, directory=True)
    workbook_path = workspace / MEMBER_NAME
    writer: pq.ParquetWriter | None = None
    batch_size = PARQUET_ROW_GROUP_SIZE
    columns: dict[str, list[Any]] = {name: [] for name in EVENT_SCHEMA.names}
    row_count = 0
    numeric_description_rows = 0
    try:
        _extract_workbook(raw_directory(data_root) / ARCHIVE_NAME, workbook_path)
        writer = pq.ParquetWriter(
            parquet_path,
            EVENT_SCHEMA,
            compression=PARQUET_COMPRESSION,
            use_dictionary=PARQUET_USE_DICTIONARY,
            write_statistics=PARQUET_WRITE_STATISTICS,
            version=PARQUET_FORMAT_VERSION,
        )
        _chmod_private(parquet_path)
        for event, numeric_description in iter_workbook_events(workbook_path):
            for name in EVENT_SCHEMA.names:
                columns[name].append(event[name])
            row_count += 1
            numeric_description_rows += int(numeric_description)
            if len(columns[EVENT_SCHEMA.names[0]]) >= batch_size:
                writer.write_table(pa.Table.from_pydict(columns, schema=EVENT_SCHEMA))
                columns = {name: [] for name in EVENT_SCHEMA.names}
        if columns[EVENT_SCHEMA.names[0]]:
            writer.write_table(pa.Table.from_pydict(columns, schema=EVENT_SCHEMA))
        writer.close()
        writer = None
        fsync_file(parquet_path)

        profile = _profile_parquet(parquet_path)
        _assert_dataset_and_stitch_profile(profile)
        manifest_profile = {
            key: profile[key]
            for key in EXPECTED_PROFILE
            if key != "numeric_description_rows"
        }
        manifest_profile["numeric_description_rows"] = numeric_description_rows
        _assert_profile(profile, manifest_profile)
        git_revision, git_dirty = _git_state()
        manifest = {
            "schema_version": NORMALIZED_MANIFEST_SCHEMA,
            "artifact_id": NORMALIZED_ARTIFACT_ID,
            "raw_artifact_id": RAW_ARTIFACT_ID,
            "raw_manifest_reference": (
                f"raw-artifact://{RAW_ARTIFACT_ID}/manifest.json"
            ),
            "dataset": {
                "path": "events.parquet",
                "bytes": parquet_path.stat().st_size,
                "sha256": sha256_file(parquet_path),
                "row_count": row_count,
                "pyarrow_schema": _schema_text(EVENT_SCHEMA),
            },
            "source_schema_version": EVENT_SCHEMA_VERSION,
            "normalizer_version": NORMALIZER_VERSION,
            "build": {
                "built_at_utc": utc_now(),
                "git_revision": git_revision,
                "git_dirty": git_dirty,
                "python_version": sys.version.split()[0],
                "dependency_versions": _dependency_versions(),
                "parquet_compression": PARQUET_COMPRESSION,
                "row_group_size": batch_size,
            },
            "stitch": {
                "rule_id": STITCH_RULE_ID,
                "cutoff_local": CUTOFF.isoformat(timespec="microseconds"),
                "first_sheet_input": EXPECTED_STITCH["first_sheet_input"],
                "first_sheet_retained": EXPECTED_STITCH["first_sheet_retained"],
                "second_sheet_input": EXPECTED_STITCH["second_sheet_input"],
                "second_sheet_retained": EXPECTED_STITCH["second_sheet_retained"],
                "excluded_cross_sheet_rows": EXPECTED_STITCH[
                    "excluded_cross_sheet_rows"
                ],
            },
            "time_semantics": {
                "field": "event_time_local",
                "timezone": "unknown",
                "cutoff_inclusive": True,
            },
            "profile": manifest_profile,
            "capabilities": CAPABILITIES,
        }
        manifest_path = stage / "manifest.json"
        manifest_path.write_bytes(_json_bytes(manifest))
        _chmod_private(manifest_path)
        fsync_file(manifest_path)
        fsync_directory(stage)
        atomic_publish_directory(stage, destination)
        return _artifact_summary("normalized", destination, manifest, "created")
    except Exception:
        if writer is not None:
            writer.close()
        shutil.rmtree(stage, ignore_errors=True)
        raise
    finally:
        shutil.rmtree(workspace, ignore_errors=True)


def verify_normalized_bundle(data_root: Path) -> dict[str, Any]:
    directory = normalized_directory(data_root)
    _assert_exact_bundle(directory, {"events.parquet", "manifest.json"})
    _verify_private_mode(directory, directory=True)
    parquet_path = directory / "events.parquet"
    manifest_path = directory / "manifest.json"
    _require_regular_file(parquet_path, "normalized Parquet")
    _require_regular_file(manifest_path, "normalized manifest")
    _verify_private_mode(parquet_path)
    _verify_private_mode(manifest_path)
    manifest = _load_json(manifest_path)
    _require_keys(
        manifest,
        {
            "schema_version",
            "artifact_id",
            "raw_artifact_id",
            "raw_manifest_reference",
            "dataset",
            "source_schema_version",
            "normalizer_version",
            "build",
            "stitch",
            "time_semantics",
            "profile",
            "capabilities",
        },
        "normalized manifest",
    )
    if manifest.get("schema_version") != NORMALIZED_MANIFEST_SCHEMA:
        raise IntegrityError("normalized manifest schema differs")
    if manifest.get("artifact_id") != NORMALIZED_ARTIFACT_ID:
        raise IntegrityError("normalized artifact ID differs")
    if manifest.get("raw_artifact_id") != RAW_ARTIFACT_ID:
        raise IntegrityError("normalized raw artifact ID differs")
    if manifest.get("raw_manifest_reference") != (
        f"raw-artifact://{RAW_ARTIFACT_ID}/manifest.json"
    ):
        raise IntegrityError("normalized raw manifest reference differs")
    dataset = manifest.get("dataset")
    if not isinstance(dataset, dict):
        raise IntegrityError("normalized dataset manifest is missing")
    _require_keys(
        dataset,
        {"path", "bytes", "sha256", "row_count", "pyarrow_schema"},
        "normalized dataset manifest",
    )
    if dataset.get("path") != "events.parquet":
        raise IntegrityError("normalized dataset path differs")
    if dataset.get("bytes") != parquet_path.stat().st_size:
        raise IntegrityError("normalized Parquet byte count differs")
    if dataset.get("sha256") != sha256_file(parquet_path):
        raise IntegrityError("normalized Parquet SHA-256 differs")
    try:
        parquet_schema = pq.ParquetFile(parquet_path).schema_arrow
    except Exception as exc:
        raise IntegrityError("normalized Parquet cannot be read") from exc
    if parquet_schema != EVENT_SCHEMA:
        raise IntegrityError("normalized Parquet schema differs")
    if dataset.get("pyarrow_schema") != _schema_text(EVENT_SCHEMA):
        raise IntegrityError("normalized manifest schema text differs")
    if dataset.get("row_count") != EXPECTED_DATASET_ROW_COUNT:
        raise IntegrityError("normalized manifest row count differs")
    if manifest.get("source_schema_version") != EVENT_SCHEMA_VERSION:
        raise IntegrityError("normalized source schema version differs")
    if manifest.get("normalizer_version") != NORMALIZER_VERSION:
        raise IntegrityError("normalized normalizer version differs")
    build = manifest.get("build")
    if not isinstance(build, dict):
        raise IntegrityError("normalized build manifest is missing")
    _require_keys(
        build,
        {
            "built_at_utc",
            "git_revision",
            "git_dirty",
            "python_version",
            "dependency_versions",
            "parquet_compression",
            "row_group_size",
        },
        "normalized build manifest",
    )
    _require_utc_timestamp(build.get("built_at_utc"), "built_at_utc")
    if build.get("python_version") != sys.version.split()[0]:
        raise IntegrityError("normalized build Python version differs")
    if build.get("dependency_versions") != _dependency_versions():
        raise IntegrityError("normalized dependency versions differ")
    if build.get("parquet_compression") != PARQUET_COMPRESSION or build.get(
        "row_group_size"
    ) != PARQUET_ROW_GROUP_SIZE:
        raise IntegrityError("normalized Parquet build settings differ")
    revision = build.get("git_revision")
    if revision is not None and (
        not isinstance(revision, str)
        or len(revision) != 40
        or any(character not in "0123456789abcdef" for character in revision)
    ):
        raise IntegrityError("normalized Git revision differs")
    dirty = build.get("git_dirty")
    if dirty is not None and not isinstance(dirty, bool):
        raise IntegrityError("normalized Git dirty state differs")
    if manifest.get("capabilities") != CAPABILITIES:
        raise IntegrityError("normalized capability declarations differ")
    stitch = manifest.get("stitch")
    expected_stitch = {
        "rule_id": STITCH_RULE_ID,
        "cutoff_local": CUTOFF.isoformat(timespec="microseconds"),
        "first_sheet_input": EXPECTED_STITCH["first_sheet_input"],
        "first_sheet_retained": EXPECTED_STITCH["first_sheet_retained"],
        "second_sheet_input": EXPECTED_STITCH["second_sheet_input"],
        "second_sheet_retained": EXPECTED_STITCH["second_sheet_retained"],
        "excluded_cross_sheet_rows": EXPECTED_STITCH["excluded_cross_sheet_rows"],
    }
    if stitch != expected_stitch:
        raise IntegrityError("normalized stitch contract differs")
    if manifest.get("time_semantics") != {
        "field": "event_time_local",
        "timezone": "unknown",
        "cutoff_inclusive": True,
    }:
        raise IntegrityError("normalized time semantics differ")
    profile = _profile_parquet(parquet_path)
    manifest_profile = manifest.get("profile")
    if not isinstance(manifest_profile, dict):
        raise IntegrityError("normalized profile is missing")
    expected_profile_keys = set(EXPECTED_PROFILE)
    if set(manifest_profile) != expected_profile_keys:
        raise IntegrityError("normalized profile keys differ")
    _assert_profile(profile, manifest_profile)
    _assert_dataset_and_stitch_profile(profile)
    for key in EXPECTED_PROFILE:
        if key == "numeric_description_rows":
            continue
        if manifest_profile.get(key) != profile[key]:
            raise IntegrityError(f"normalized manifest profile differs: {key}")
    return manifest


def verify(data_root: Path) -> dict[str, Any]:
    raw_manifest = verify_raw_bundle(data_root.resolve())
    normalized_manifest = verify_normalized_bundle(data_root.resolve())
    return {
        "raw": _artifact_summary(
            "raw", raw_directory(data_root.resolve()), raw_manifest, "verified"
        ),
        "normalized": _artifact_summary(
            "normalized",
            normalized_directory(data_root.resolve()),
            normalized_manifest,
            "verified",
        ),
        "statistics": normalized_manifest["profile"],
    }


def make_run_id(now: datetime | None = None) -> str:
    instant = now or datetime.now(UTC)
    stamp = instant.strftime("%Y%m%dT%H%M%S%fZ")
    return f"{stamp}-{uuid.uuid4().hex[:8]}"


def write_receipt(data_root: Path, receipt: dict[str, Any]) -> Path:
    resolved_root = data_root.resolve()
    directory = receipts_directory(resolved_root)
    ensure_private_data_tree(resolved_root, directory)
    final = directory / f"{receipt['run_id']}.json"
    temporary = directory / f".{receipt['run_id']}.{uuid.uuid4().hex}.part"
    try:
        with temporary.open("xb") as handle:
            _chmod_private(temporary)
            handle.write(_json_bytes(receipt))
            handle.flush()
            os.fsync(handle.fileno())
        _publish_noreplace(temporary, final, directory=False)
        return final
    except Exception:
        temporary.unlink(missing_ok=True)
        raise
