from __future__ import annotations

from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook

from productquant import uci
from productquant.errors import IntegrityError

from conftest import HEADERS


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        (123, "123"),
        (123.0, "123"),
        (" 00123 ", " 00123 "),
        ("Mixed-Case", "Mixed-Case"),
    ],
)
def test_identifier_conversion_is_lossless(value: object, expected: str) -> None:
    assert uci._as_identifier(value, "identifier") == expected


@pytest.mark.parametrize("value", [None, True, 1.5, Decimal("1.5"), object()])
def test_identifier_conversion_rejects_unsupported_values(value: object) -> None:
    with pytest.raises(IntegrityError, match="identifier"):
        uci._as_identifier(value, "identifier")


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        (None, (None, False)),
        ("  observed text  ", ("  observed text  ", False)),
        (12345, ("12345", True)),
        (12345.0, ("12345", True)),
    ],
)
def test_description_conversion_preserves_text_and_marks_numeric_cells(
    value: object, expected: tuple[str | None, bool]
) -> None:
    assert uci._as_description(value) == expected


@pytest.mark.parametrize("value", [1.5, True, Decimal("2.5"), object()])
def test_description_conversion_rejects_lossy_values(value: object) -> None:
    with pytest.raises(IntegrityError, match="Description"):
        uci._as_description(value)


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        (0, 0),
        (-80_995, -80_995),
        (80_995, 80_995),
        (-(1 << 63), -(1 << 63)),
        ((1 << 63) - 1, (1 << 63) - 1),
        (2.0, 2),
    ],
)
def test_quantity_conversion_preserves_signed_integral_values(
    value: object, expected: int
) -> None:
    assert uci._as_quantity(value) == expected


@pytest.mark.parametrize("value", [None, True, 1.5, Decimal("1"), "1"])
def test_quantity_conversion_rejects_non_excel_integers(value: object) -> None:
    with pytest.raises(IntegrityError, match="Quantity"):
        uci._as_quantity(value)


@pytest.mark.parametrize("value", [-(1 << 63) - 1, 1 << 63])
def test_quantity_conversion_rejects_signed_int64_overflow(value: int) -> None:
    with pytest.raises(IntegrityError, match="Quantity does not fit int64"):
        uci._as_quantity(value)


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        (1, Decimal("1.000")),
        (1.2, Decimal("1.200")),
        (-1.125, Decimal("-1.125")),
        (Decimal("38970.000"), Decimal("38970.000")),
    ],
)
def test_price_conversion_preserves_exact_three_decimal_contract(
    value: object, expected: Decimal
) -> None:
    assert uci._as_price(value) == expected


@pytest.mark.parametrize(
    "value",
    [None, True, "1.000", 1.0001, Decimal("1234567890123456.789")],
)
def test_price_conversion_fails_closed_on_invalid_or_lossy_values(
    value: object,
) -> None:
    with pytest.raises(IntegrityError, match="Price"):
        uci._as_price(value)


def test_row_to_event_preserves_provenance_nulls_whitespace_and_cancellation(
    synthetic_source: object,
) -> None:
    row = (
        "C001 ",
        " SKU-1 ",
        12345,
        -2,
        datetime(2010, 11, 30, 12),
        -1.125,
        None,
        " United Kingdom ",
    )

    event, numeric_description = uci._row_to_event("Year 2009-2010", 17, row)

    assert event == {
        "event_id": f"{uci.RAW_ARTIFACT_ID}:year-2009-2010:17",
        "raw_artifact_id": uci.RAW_ARTIFACT_ID,
        "source_id": "uci-online-retail-ii",
        "source_member": "online_retail_II.xlsx",
        "source_sheet": "Year 2009-2010",
        "source_row_number": 17,
        "event_time_local": datetime(2010, 11, 30, 12),
        "source_invoice_id": "C001 ",
        "source_product_code": " SKU-1 ",
        "description_observed": "12345",
        "quantity": -2,
        "unit_price": Decimal("-1.125"),
        "currency_code": "GBP",
        "customer_reference": None,
        "country_observed": " United Kingdom ",
        "invoice_is_cancellation": True,
        "schema_version": "transaction_event.v1",
    }
    assert numeric_description is True


@pytest.mark.parametrize(
    "row",
    [
        (None, 1, "d", 1, datetime(2010, 1, 1), 1, None, "UK"),
        (1, None, "d", 1, datetime(2010, 1, 1), 1, None, "UK"),
        (1, 1, "d", 1.5, datetime(2010, 1, 1), 1, None, "UK"),
        (1, 1, "d", 1, "2010-01-01", 1, None, "UK"),
        (1, 1, "d", 1, datetime(2010, 1, 1, tzinfo=UTC), 1, None, "UK"),
        (1, 1, "d", 1, datetime(2010, 1, 1), 1.0001, None, "UK"),
        (1, 1, "d", 1, datetime(2010, 1, 1), 1, None, None),
    ],
)
def test_row_to_event_fails_closed_on_required_type_drift(
    row: tuple[object, ...],
) -> None:
    with pytest.raises(IntegrityError):
        uci._row_to_event("Year 2009-2010", 2, row)


def _write_workbook(
    path: Path,
    *,
    sheets: tuple[str, ...] = ("Year 2009-2010", "Year 2010-2011"),
    headers: tuple[str, ...] = HEADERS,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in sheets:
        sheet = workbook.create_sheet(name)
        sheet.append(headers)
    workbook.save(path)
    workbook.close()


def test_iter_workbook_events_rejects_sheet_set_or_order_drift(tmp_path: Path) -> None:
    path = tmp_path / "wrong-sheets.xlsx"
    _write_workbook(path, sheets=("Year 2010-2011", "Year 2009-2010"))

    with pytest.raises(IntegrityError, match="sheet set/order"):
        list(uci.iter_workbook_events(path))


def test_iter_workbook_events_rejects_header_drift(tmp_path: Path) -> None:
    path = tmp_path / "wrong-header.xlsx"
    changed = (*HEADERS[:-1], "CountryName")
    _write_workbook(path, headers=changed)

    with pytest.raises(IntegrityError, match="header differs"):
        list(uci.iter_workbook_events(path))


@pytest.mark.parametrize("cell_value", ["=1+1", "#DIV/0!"])
def test_iter_workbook_events_rejects_formula_and_error_cells_before_coercion(
    tmp_path: Path,
    cell_value: str,
) -> None:
    path = tmp_path / "unsupported-cell.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "Year 2009-2010"
    second = workbook.create_sheet("Year 2010-2011")
    first.append(HEADERS)
    second.append(HEADERS)
    first.append(
        (
            "1",
            "SKU",
            cell_value,
            1,
            datetime(2010, 1, 1),
            1,
            None,
            "UK",
        )
    )
    workbook.save(path)
    workbook.close()

    with pytest.raises(IntegrityError, match="formula or error cell"):
        list(uci.iter_workbook_events(path))


def test_iter_workbook_events_rejects_corrupt_workbook(tmp_path: Path) -> None:
    path = tmp_path / "corrupt.xlsx"
    path.write_bytes(b"not an xlsx")

    with pytest.raises(IntegrityError, match="cannot be opened"):
        list(uci.iter_workbook_events(path))
